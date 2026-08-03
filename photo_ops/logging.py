import csv
import io
import os
import threading
import uuid
from datetime import datetime

# Define PathLike locally to avoid circular imports
PathLike = str | os.PathLike

_log_lock = threading.Lock()

import atexit

# ── xlsxwriter workbook cache ─────────────────────────────────────────────────
# Each log_file path maps to an open xlsxwriter Workbook + per-sheet writers.
# xlsxwriter does NOT support reading back existing files, so we always create
# fresh workbooks and accumulate rows in memory until flush / process exit.
_xw_cache: dict[str, dict] = {}  # path → {"wb": wb, "sheets": {name: ws}, "row": {name: int}}

_THUMB_W = 120  # pixels – before/after thumbnail width
_THUMB_H = 120  # pixels – before/after thumbnail height
_ROW_H = 92  # Excel row height (points) when thumbnails are shown
_COL_W = 16  # Excel column width (chars) for thumbnail columns

_HEADERS_PLAIN = ["Timestamp", "File", "Status", "Actions / Reason"]
_HEADERS_THUMB = ["Timestamp", "File", "Status", "Actions / Reason", "Before", "After"]


def _make_thumb_bytes(img_path: str) -> bytes | None:
    """Return JPEG thumbnail bytes using the Rust engine; None on any failure."""
    if not img_path or not os.path.exists(img_path):
        return None
    try:
        import photo_ops

        buf = photo_ops.resize(img_path, None, _THUMB_W, _THUMB_H, True)
        if isinstance(buf, (bytes, bytearray)):
            return bytes(buf)
        # Fallback: resize to a tmp file and read it back
        tmp = os.path.join(os.path.dirname(img_path), f"_t_{uuid.uuid4().hex}.jpg")
        try:
            photo_ops.resize(img_path, tmp, _THUMB_W, _THUMB_H, True)
            if os.path.exists(tmp):
                with open(tmp, "rb") as fh:
                    return fh.read()
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)
    except Exception:
        return None


def _get_xw_sheet(log_file: str, sheet_name: str, thumbnail: bool):
    """Return (wb_entry, worksheet, next_row) — creating as needed."""
    if log_file not in _xw_cache:
        try:
            import xlsxwriter
        except ImportError:
            return None, None, None

        wb = xlsxwriter.Workbook(log_file, {"constant_memory": False})
        headers = _HEADERS_THUMB if thumbnail else _HEADERS_PLAIN

        # Pre-create the three standard sheets so tab order is predictable
        sheets = {}
        rows = {}
        for name in ("Success", "Failed", "Subjects"):
            ws = wb.add_worksheet(name)
            # Header row formatting
            hdr_fmt = wb.add_format(
                {"bold": True, "bg_color": "#2D3250", "font_color": "#FFFFFF", "border": 1}
            )
            for col, hdr in enumerate(headers):
                ws.write(0, col, hdr, hdr_fmt)
            ws.set_row(0, 18)
            ws.set_column(0, 0, 20)  # Timestamp
            ws.set_column(1, 1, 30)  # File
            ws.set_column(2, 2, 10)  # Status
            ws.set_column(3, 3, 40)  # Actions / Reason
            if thumbnail:
                ws.set_column(4, 4, _COL_W)  # Before
                ws.set_column(5, 5, _COL_W)  # After
            sheets[name] = ws
            rows[name] = 1  # next available row (0 = header)

        _xw_cache[log_file] = {"wb": wb, "sheets": sheets, "rows": rows, "thumbnail": thumbnail}

    entry = _xw_cache[log_file]
    wb = entry["wb"]

    if sheet_name not in entry["sheets"]:
        ws = wb.add_worksheet(sheet_name)
        headers = _HEADERS_THUMB if entry["thumbnail"] else _HEADERS_PLAIN
        hdr_fmt = wb.add_format(
            {"bold": True, "bg_color": "#2D3250", "font_color": "#FFFFFF", "border": 1}
        )
        for col, hdr in enumerate(headers):
            ws.write(0, col, hdr, hdr_fmt)
        ws.set_column(0, 3, 20)
        if entry["thumbnail"]:
            ws.set_column(4, 5, _COL_W)
        entry["sheets"][sheet_name] = ws
        entry["rows"][sheet_name] = 1

    return entry, entry["sheets"][sheet_name], entry["rows"][sheet_name]


def _flush_xw(log_file: str):
    """Close and write the xlsxwriter workbook to disk."""
    entry = _xw_cache.pop(log_file, None)
    if entry:
        try:
            entry["wb"].close()
        except Exception:
            pass


def _flush_all_xw():
    for path in list(_xw_cache.keys()):
        _flush_xw(path)


atexit.register(_flush_all_xw)


def log_advanced(
    file_path: PathLike,
    status: str,
    actions: str,
    reason: str = "",
    log_file: str = "log.xlsx",
    split_csv: bool = True,
    thumbnail: bool = False,
    before_path: PathLike | None = None,
) -> None:
    """
    Write one audit row to *log_file*.

    Parameters
    ----------
    file_path   : Path to the **output** (after) file.
    status      : ``"success"`` | ``"failed"`` | ``"subject"``
    actions     : Comma-joined list of ops applied (used when status=success).
    reason      : Error message (used when status=failed).
    log_file    : Destination path.  Extension determines format:
                  ``.xlsx`` → xlsxwriter with optional embedded thumbnails,
                  ``.csv``  → plain CSV,
                  other     → TSV.
    split_csv   : When True and format is CSV, write separate ``_success``/
                  ``_failed`` files.
    thumbnail   : Embed 120×120 before/after thumbnails (XLSX only).
    before_path : Path to the **input** (before) file.  Required for the
                  "Before" thumbnail column; ignored for CSV/TSV.
    """
    with _log_lock:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ext = os.path.splitext(log_file)[1].lower()
        filename = os.path.basename(str(file_path))
        details = actions if status == "success" else reason

        try:
            # ── XLSX ──────────────────────────────────────────────────────────
            if ext == ".xlsx":
                if thumbnail and ext == ".csv":
                    # Safety net: shouldn't reach here, but just in case
                    log_file = log_file.replace(".csv", ".xlsx")
                    ext = ".xlsx"

                # Try xlsxwriter path first
                entry, ws, row_idx = _get_xw_sheet(
                    log_file,
                    "Success"
                    if status == "success"
                    else ("Failed" if status == "failed" else "Subjects"),
                    thumbnail,
                )

                if ws is not None:
                    sheet_name = (
                        "Success"
                        if status == "success"
                        else ("Failed" if status == "failed" else "Subjects")
                    )

                    # Cell formats
                    wb = entry["wb"]
                    normal = wb.add_format({"border": 1, "valign": "vcenter"})
                    green = wb.add_format(
                        {"border": 1, "valign": "vcenter", "font_color": "#1A7431", "bold": True}
                    )
                    red = wb.add_format(
                        {"border": 1, "valign": "vcenter", "font_color": "#C0392B", "bold": True}
                    )
                    s_fmt = (
                        green if status == "success" else (red if status == "failed" else normal)
                    )

                    ws.write(row_idx, 0, timestamp, normal)
                    ws.write(row_idx, 1, filename, normal)
                    ws.write(row_idx, 2, status, s_fmt)
                    ws.write(row_idx, 3, details, normal)

                    if thumbnail:
                        ws.set_row(row_idx, _ROW_H)

                        # Before thumbnail
                        before_bytes = _make_thumb_bytes(str(before_path)) if before_path else None
                        if before_bytes:
                            ws.insert_image(
                                row_idx,
                                4,
                                f"before_{filename}",
                                {
                                    "image_data": io.BytesIO(before_bytes),
                                    "x_scale": 1,
                                    "y_scale": 1,
                                    "object_position": 1,
                                },
                            )

                        # After thumbnail
                        after_bytes = _make_thumb_bytes(str(file_path))
                        if after_bytes:
                            ws.insert_image(
                                row_idx,
                                5,
                                f"after_{filename}",
                                {
                                    "image_data": io.BytesIO(after_bytes),
                                    "x_scale": 1,
                                    "y_scale": 1,
                                    "object_position": 1,
                                },
                            )

                    entry["rows"][sheet_name] = row_idx + 1

                    # Periodic flush every 100 rows to guard against crashes
                    total = sum(entry["rows"].values())
                    if total % 100 == 0:
                        # xlsxwriter can't mid-stream save; we skip intermediate saves
                        # but flush is guaranteed at process exit via atexit hook.
                        pass

                else:
                    # xlsxwriter not available — fall back to openpyxl (no thumbnails)
                    try:
                        import openpyxl
                        from openpyxl import Workbook

                        if os.path.exists(log_file):
                            wb2 = openpyxl.load_workbook(log_file)
                        else:
                            wb2 = Workbook()
                            wb2.active.title = "Success"
                            wb2.create_sheet("Failed")
                            wb2["Success"].append(_HEADERS_PLAIN)
                            wb2["Failed"].append(_HEADERS_PLAIN)

                        sheet_name2 = (
                            "Success"
                            if status == "success"
                            else ("Failed" if status == "failed" else "Subjects")
                        )
                        if sheet_name2 not in wb2.sheetnames:
                            wb2.create_sheet(sheet_name2)
                            wb2[sheet_name2].append(_HEADERS_PLAIN)

                        wb2[sheet_name2].append([timestamp, filename, status, details])
                        wb2.save(log_file)
                    except ImportError:
                        # Last resort: CSV
                        _write_csv(
                            log_file.replace(".xlsx", ".csv"),
                            split_csv,
                            status,
                            timestamp,
                            filename,
                            details,
                            str(file_path),
                        )

            # ── CSV ───────────────────────────────────────────────────────────
            elif ext == ".csv":
                if thumbnail:
                    print("[WARNING] Thumbnails are not supported in CSV logs. Use XLSX format.")
                _write_csv(
                    log_file, split_csv, status, timestamp, filename, details, str(file_path)
                )

            # ── TSV / other ───────────────────────────────────────────────────
            else:
                write_header = not os.path.exists(log_file)
                with open(log_file, "a", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f, delimiter="\t")
                    if write_header:
                        writer.writerow(_HEADERS_PLAIN)
                    writer.writerow([timestamp, filename, status, details])

        except Exception as e:
            print(f"Logging failed: {e}")


def _write_csv(
    log_file: str,
    split_csv: bool,
    status: str,
    timestamp: str,
    filename: str,
    details: str,
    file_path: str,
):
    if split_csv:
        base, _ = os.path.splitext(log_file)
        log_file = f"{base}_{status}.csv"
    write_header = not os.path.exists(log_file)
    with open(log_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(["Timestamp", "File", "Status", "Details", "Link"])
        link = file_path if status == "failed" else ""
        writer.writerow([timestamp, filename, status, details, link])
